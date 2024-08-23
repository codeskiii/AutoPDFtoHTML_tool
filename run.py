import argparse
import os
from openpyxl import load_workbook
import requests
from bs4 import BeautifulSoup
from spire.pdf.common import *
from spire.pdf import *
from spire.pdf import PdfDocument, FileFormat
import logging
from pathlib import Path
from typing import List, Tuple
import re

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

class DataDownloader:
    def __init__(self, base_urls_file: str, to_visit_urls_file: str) -> None:
        os.makedirs('downloaded', exist_ok=True)

        with open(base_urls_file, 'r') as f:
            self.base_url = f.read().strip()

        with open(to_visit_urls_file, 'r') as f:
            urls_raw = f.readlines()
        self.urls_to_visit = [self.base_url + line.strip() for line in urls_raw]

    def _get_req_elements(self, path: str) -> list:
        workbook = load_workbook(path)
        sheet = workbook.active
        headers = [cell.value for cell in sheet[1]]  # Get headers from the first row
        return [dict(zip(headers, row)) for row in sheet.iter_rows(values_only=True, min_row=2)]

    def download_save_data(
        self,
        to_look_for_file_path: str,
        auth: Tuple[str, str],
        splitter: str,
        types_to_look_file_path: str,
        columns_of_to_look_for: List[str]
    ) -> None:
        def load_types_to_look(file_path: str) -> List[str]:
            with open(file_path, 'r') as file:
                return [line.strip() for line in file.readlines()]

        def download_pdf(pdf_url: str, file_path: str) -> None:
            try:
                pdf_response = requests.get(pdf_url, auth=auth)
                pdf_response.raise_for_status()
                with open(file_path, "wb") as f:
                    f.write(pdf_response.content)
                print(f"Saved PDF to {file_path}")
            except requests.RequestException as e:
                print(f"Failed to download PDF: {e}")

        def ensure_directory_exists(directory: str) -> None:
            if not os.path.exists(directory):
                os.makedirs(directory)

        # Load elements to look for
        to_look_for_stuff = self._get_req_elements(to_look_for_file_path)
        del to_look_for_stuff[0]

        # Load types to look for
        types_to_look = load_types_to_look(types_to_look_file_path)

        # Ensure the directory exists
        ensure_directory_exists("downloaded")


        for url in self.urls_to_visit:
            print(f"Downloading data from: {url}")
            try:
                response = requests.get(url, auth=auth)
                response.raise_for_status()
            except requests.RequestException as e:
                print(f"Failed to fetch data from {url}: {e}")
                continue

            soup = BeautifulSoup(response.content, 'html.parser')
            for tr in soup.find_all('tr'):
                #parts = [td.get_text(strip=True) for td in tr.find_all('td')]
                parts = []
                for td in tr.find_all('td'):
                    text = td.get_text(strip=True)
                    texts = re.split(r'[ _]', text)
                    texts = [e.upper() for e in texts]
                    text = '_'.join(texts)
                    parts.append(text)

                #print(parts)

                if parts:
                    for item in to_look_for_stuff:
                        #print(proccessed_to_look_for)

                        proccesed_c = []
                        for collumn in columns_of_to_look_for:
                            var = item.get(collumn)
                            if var:
                                var = re.split(r'[ _]', var)
                                var = [e.upper() for e in var]
                                var = '_'.join(var)
                                proccesed_c.append(var)

                        #print('=============')
                        #print(proccesed_c)
                        #print(parts)

                        if all( collumn in parts for collumn in proccesed_c):
                            links = tr.find_all("a")
                            for link in links:
                                link_text = link.get_text(strip=True)
                                if link_text.split(splitter)[0] in types_to_look:
                                    try:
                                        response = requests.get(link['href'], auth=auth)
                                        response.raise_for_status()
                                    except requests.RequestException as e:
                                        print(f"Failed to fetch data from link {link['href']}: {e}")
                                        continue

                                    soup = BeautifulSoup(response.content, 'html.parser')
                                    tds = soup.find_all('td')
                                    for td in tds:
                                        ass = td.find_all('a')
                                        for e in ass:
                                            if e.text == 'Download':
                                                pdf_url = e.get('href')
                                                pdf_filename = f"downloaded/{item.get('Id')}_{link_text.split(splitter)[0]}.pdf"
                                                print(f"Downloading from: {pdf_url}")
                                                download_pdf(pdf_url, pdf_filename)


class DataProcessor:
    def __init__(self) -> None:
        self.files = [f for f in os.listdir('downloaded') if f.endswith('.pdf')]

    def rewrite_files(self, out_dir: str) -> None:
        out_dir_path = out_dir

        for index, file_name in enumerate(self.files):
            logging.info(f"Processing file: {file_name}")
            pdf_path = 'downloaded' + '//' + file_name

            try:
                doc = PdfDocument()
                logging.info(f"Loading PDF from path: {pdf_path}")
                doc.LoadFromFile(pdf_path)

                logging.info(f"Setting opttions: {pdf_path}")
                convert_options = doc.ConvertOptions
                convert_options.SetPdfToHtmlOptions(True, True, 1, True)

                # Extract filename without extension for output
                output_filename = os.path.splitext(str(file_name))[0]
                logging.info(f"Making dir for: {output_filename}")
                out_file_path = out_dir_path + '//' + f"{output_filename}.html"

                doc.SaveToFile(out_file_path, FileFormat.HTML)
                logging.info(f"Converted {file_name} to HTML and saved to {out_file_path}")

            except FileNotFoundError:
                logging.error(f"File not found: {pdf_path}")
            except TypeError:
                logging.error(f"Invalid PDF format: {pdf_path}")
            except Exception as e:
                logging.error(f"Error processing {file_name}: {e}")
            finally:
                doc.Dispose()

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Auto Docs")
    parser.add_argument('base_link_file', type=str, help='File with base link')
    parser.add_argument('to_visit_links_file', type=str, help='File with links to visit')
    parser.add_argument('to_look_for', type=str, help='File with items to look for')
    parser.add_argument('types', type=str, nargs='+', help='Types to look for (space-separated)')
    parser.add_argument('--type_separator', type=str, default='-', help='Types separator, e.g., "-" in "isin-fx"')
    parser.add_argument('--output_dir', type=str, help='Output directory for converted files')
    parser.add_argument('--username', type=str, default='user', help='Username for authentication')
    parser.add_argument('--password', type=str, default='password', help='Password for authentication')
    parser.add_argument('--columns', type=str, nargs='+', help='Columns to look for (space-separated)')

    args = parser.parse_args()

    # Initialize DataDownloader and download PDF files
    downloader = DataDownloader(args.base_link_file, args.to_visit_links_file)
    downloader.download_save_data(args.to_look_for, 
                                  (args.username, args.password), args.type_separator, args.types[0], args.columns)

    # Initialize DataProcessor and convert downloaded PDF files
    if args.output_dir:
        processor = DataProcessor()
        processor.rewrite_files(args.output_dir)
